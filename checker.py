import requests
from requests import Session
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import threading
import random

FILE_PATH = 'Data File.xlsx'
wb = load_workbook(FILE_PATH)

# creating a session object
session = Session()
HEADERS_LIST = [
    'Mozilla/5.0 (Windows; U; Windows NT 6.1; x64; fr; rv:1.9.2.13) Gecko/20101203 Firebird/3.6.13',
    'Mozilla/5.0 (compatible, MSIE 11, Windows NT 6.3; Trident/7.0; rv:11.0) like Gecko',
    'Mozilla/5.0 (Windows; U; Windows NT 6.1; rv:2.2) Gecko/20110201',
    'Opera/9.80 (X11; Linux i686; Ubuntu/14.10) Presto/2.12.388 Version/12.16',
    'Mozilla/5.0 (Windows NT 5.2; RW; rv:7.0a1) Gecko/20091211 SeaMonkey/9.23a1pre'
]

HEADER = {'User-Agent': random.choice(HEADERS_LIST), 'X-Requested-With': 'XMLHttpRequest'}
session.headers.update(HEADER)


# takes an url and returns its status code and final redirected url
def get_response_code(url):
    url = 'https://' + url if not url.startswith('http') else url
    error = None
    try:
        response = session.get(url)
        final_url = response.url
        status_code = response.status_code
        if response.history:
            status_code = response.history[0].status_code
    except requests.exceptions.SSLError:
        response = session.get(url, verify=False)
        final_url = response.url
        status_code = response.status_code
    except requests.exceptions.ConnectionError as err:
        final_url = url 
        status_code = 'ERROR'
        error = str(err)
    return (url, final_url, status_code, error)


def customize_excel_sheet():
    global wb
    output = wb.create_sheet('Output') if 'Output' not in wb.sheetnames else wb['Output']
    
    font = Font(color="000000", bold=True)
    bg_color = PatternFill(fgColor='E8E8E8', fill_type='solid')

    # editing the output sheet
    output_column = zip(('A',  'B', 'C', 'D'), ('URL', 'Final URL', 'HTTP Response', 'Error Details'))
    for col, value in output_column:
        cell = output[f'{col}1']
        cell.value = value
        cell.font = font
        cell.fill = bg_color
        output.freeze_panes = cell

        # fixing the column width
        output.column_dimensions[col].width = 20


# Generates the input links
def generate_input_urls():
    global wb
    inputs = wb['Input']
    for row in range(2, inputs.max_row + 1):
        # generates the links one by one
        if value := inputs[f"A{row}"].value:
            yield value


def insert_data_to_excel():
    global wb

    # gets the data and inserts it to excel
    def save(url):
        data = get_response_code(url)
        wb['Output'].append(data)
        print(data)

    # applying threading to reduce execution time
    threads = []
    for url in generate_input_urls():
        thread = threading.Thread(target=save, args=[url])
        thread.start()
        threads.append(thread)
        
    # joining threads so that functions after this run after the completion of the threads
    for thread in threads:
        thread.join()

def main():
    customize_excel_sheet()
    insert_data_to_excel()
    wb.save(FILE_PATH)


main()
